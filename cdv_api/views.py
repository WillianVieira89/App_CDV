import json
import logging
import datetime
from collections import defaultdict

import openpyxl
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.db import transaction
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import TemplateDoesNotExist
from django.utils import timezone
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from .models import Estacao, Transmissor, Receptor

logger = logging.getLogger(__name__)


# =========================
# UTILITÁRIOS
# =========================

def safe_float(value):
    try:
        return float(value) if value is not None else None
    except (ValueError, TypeError):
        return None


def safe_int(value):
    try:
        return int(value) if value is not None else None
    except (ValueError, TypeError):
        return None


def _norm_manutencao(valor):
    if not valor:
        return "preventiva"

    v = str(valor).strip().lower()
    mapa = {
        "preventiva": "preventiva",
        "corretiva": "corretiva",
        "check-list": "checklist",
        "checklist": "checklist",
    }
    return mapa.get(v, "preventiva")


def _pick_temp(d):
    return safe_float(d.get("temp_celsius", d.get("temperatura_local")))


def relacao_para_float(relacao_str):
    if not relacao_str:
        return None

    try:
        return float(str(relacao_str).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def ordenar_estacoes_linha(estacoes_qs):
    ordem_linha = [
        "Capão Redondo",
        "Campo Limpo",
        "Vila das Belezas",
        "Giovanni Gronchi",
        "Santo Amaro",
        "Largo Treze",
        "Adolfo Pinheiro",
        "Alto da Boa Vista",
        "Borba Gato",
        "Brooklin",
        "Campo Belo",
        "Eucaliptos",
        "Moema",
        "AACD Servidor",
        "Hospital São Paulo",
        "Santa Cruz",
        "Chácara Klabin",
    ]

    ordem_idx = {nome: i for i, nome in enumerate(ordem_linha)}

    estacoes = list(estacoes_qs)
    estacoes.sort(key=lambda e: ordem_idx.get(e.nome, 999))
    return estacoes


def detectar_degradacao_faixa(receptores_queryset, qtd_leituras=3):
    """
    Detecta degradação quando a tendência sai da faixa normal (60% a 80%).

    Regras:
    - usa as últimas 3 leituras por circuito
    - se estiver em queda contínua e a última < 60 -> degradação negativa
    - se estiver em subida contínua e a última > 80 -> degradação positiva
    """
    historico_por_circuito = defaultdict(list)

    receptores_ordenados = receptores_queryset.order_by("num_circuito", "-data_manutencao", "-id")

    for r in receptores_ordenados:
        valor = relacao_para_float(r.relacao)
        if valor is None:
            continue

        circuito = (r.num_circuito or "").strip().upper()

        if len(historico_por_circuito[circuito]) < qtd_leituras:
            historico_por_circuito[circuito].append({
                "data": r.data_manutencao,
                "relacao": valor,
            })

    circuitos_degradados = []

    for circuito, leituras in historico_por_circuito.items():
        if len(leituras) < qtd_leituras:
            continue

        leituras = list(reversed(leituras))  # mais antiga -> mais recente
        valores = [item["relacao"] for item in leituras]

        em_queda = all(valores[i] > valores[i + 1] for i in range(len(valores) - 1))
        em_subida = all(valores[i] < valores[i + 1] for i in range(len(valores) - 1))

        ultima = valores[-1]
        primeira = valores[0]

        if em_queda and ultima < 60:
            circuitos_degradados.append({
                "circuito": circuito,
                "leituras": [round(v, 2) for v in valores],
                "variacao_total": round(ultima - primeira, 2),
                "ultima_relacao": round(ultima, 2),
                "tipo_degradacao": "Negativa",
                "status": "Abaixo de 60%",
            })

        elif em_subida and ultima > 80:
            circuitos_degradados.append({
                "circuito": circuito,
                "leituras": [round(v, 2) for v in valores],
                "variacao_total": round(ultima - primeira, 2),
                "ultima_relacao": round(ultima, 2),
                "tipo_degradacao": "Positiva",
                "status": "Acima de 80%",
            })

    circuitos_degradados.sort(
        key=lambda x: (
            x["tipo_degradacao"] != "Negativa",
            x["ultima_relacao"]
        )
    )

    return circuitos_degradados


def classificar_relacao(valor):
    if valor is None:
        return "Sem dado", ""

    if valor < 60:
        return "Abaixo de 60%", "relacao-baixa"

    if valor > 80:
        return "Acima de 80%", "relacao-alta"

    return "Entre 60% e 80%", "relacao-normal"


def identificar_via(circuito):
    circuito = (circuito or "").strip().upper()

    if circuito.startswith("1"):
        return "Via 01"

    if circuito.startswith("2"):
        return "Via 02"

    return "Não definida"

def calcular_radar_saude(relacao, temperatura):
    if relacao is None:
        return {
            "score": 0,
            "status": "Sem dados",
            "cor": "vermelho",
            "tipo": "Indefinido"
        }

    if relacao >= 100:
        tipo = "Falsa ocupação"
        score = 0
    elif relacao > 80:
        tipo = "Sensível"
        score = 40
    elif relacao >= 60:
        tipo = "Normal"
        score = 100
    else:
        tipo = "Falsa desocupação"
        score = 20

    if temperatura is not None:
        if temperatura > 50:
            score -= 10
        elif temperatura < 10:
            score -= 5

    score = max(0, score)

    if score >= 80:
        cor = "verde"
        status = "Saudável"
    elif score >= 40:
        cor = "amarelo"
        status = "Atenção"
    else:
        cor = "vermelho"
        status = "Crítico"

    return {
        "score": score,
        "status": status,
        "cor": cor,
        "tipo": tipo
    }
    
# =========================
# PÁGINAS PRINCIPAIS
# =========================

@login_required
def index(request):
    lista_de_estacoes = ordenar_estacoes_linha(Estacao.objects.all())
    context = {
        "lista_de_estacoes": lista_de_estacoes,
        "selected_estacao_id": request.GET.get("estacao_id"),
    }
    try:
        return render(request, "cdv_api/index.html", context)
    except TemplateDoesNotExist:
        logger.exception("Template cdv_api/index.html ausente")
        return HttpResponse("<h1>Home OK</h1>")


@login_required
def home(request):
    try:
        return render(request, "cdv_api/home.html", {})
    except TemplateDoesNotExist:
        logger.exception("Falha ao renderizar home()")
        return render(request, "cdv_api/erro_generico.html", status=500)


def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = authenticate(
                username=form.cleaned_data.get("username"),
                password=form.cleaned_data.get("password"),
            )
            if user is not None:
                login(request, user)
                return redirect("index")

            form.add_error(None, "Nome de usuário ou senha incorretos.")

        return render(request, "cdv_api/login.html", {"form": form})

    form = AuthenticationForm()
    return render(request, "cdv_api/login.html", {"form": form})


@login_required
def registrar_cdv(request):
    estacao_param = request.GET.get("estacao")
    if not estacao_param:
        return redirect("index")

    try:
        try:
            estacao = Estacao.objects.get(id=int(estacao_param))
        except (ValueError, Estacao.DoesNotExist):
            estacao = Estacao.objects.get(nome=estacao_param)
    except Estacao.DoesNotExist:
        return redirect("index")

    tx_codes = Transmissor.objects.filter(estacao=estacao).values_list("num_circuito", flat=True)
    rx_codes = Receptor.objects.filter(estacao=estacao).values_list("num_circuito", flat=True)
    circuitos = sorted({(c or "").strip() for c in list(tx_codes) + list(rx_codes) if c})

    context = {
        "estacao_nome": estacao.nome,
        "estacao_id_current": estacao.id,
        "circuitos": circuitos,
    }
    return render(request, "cdv_api/registrar_cdv.html", context)


@login_required
def gerar_relatorio_excel_page(request):
    lista_de_estacoes = ordenar_estacoes_linha(Estacao.objects.all())

    selected_estacao_id = request.GET.get("estacao_id", "")
    circuito_filtro = request.GET.get("circuito_filtro", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")
    tipo_manutencao = request.GET.get("tipo_manutencao", "")

    estacao_nome = None
    if selected_estacao_id:
        try:
            estacao_nome = Estacao.objects.get(id=selected_estacao_id).nome
        except Estacao.DoesNotExist:
            selected_estacao_id = ""
            estacao_nome = None

    estacoes_mapa = []
    for est in lista_de_estacoes:
        estacoes_mapa.append({
            "id": est.id,
            "nome": est.nome,
            "status": "normal",
        })

    return render(
        request,
        "cdv_api/gerar_relatorio_excel.html",
        {
            "lista_de_estacoes": lista_de_estacoes,
            "selected_estacao_id": str(selected_estacao_id) if selected_estacao_id else "",
            "estacao_nome": estacao_nome,
            "circuito_filtro": circuito_filtro,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "tipo_manutencao": tipo_manutencao,
            "estacoes_mapa": estacoes_mapa,
        },
    )

@login_required
def radar_saude(request):
    receptores = Receptor.objects.all().order_by("num_circuito", "-data_manutencao")

    # Último registro por circuito
    ultimo_por_circuito = {}

    for r in receptores:
        chave = r.num_circuito
        if chave not in ultimo_por_circuito:
            ultimo_por_circuito[chave] = r

    radar_lista = []

    for circuito, r in ultimo_por_circuito.items():
        rel = None
        if r.relacao:
            try:
                rel = float(r.relacao.replace("%", "").replace(",", "."))
            except:
                pass

        # calcula o radar para este receptor
        radar = calcular_radar_saude(rel, r.temp_celsius)

        radar_lista.append({
            "estacao": r.estacao.nome if r.estacao else "-",
            "circuito": circuito,
            "rx": r.num_receptor,
            "relacao": rel,
            "temperatura": r.temp_celsius,
            "score": radar["score"],
            "status": radar["status"],
            "cor": radar["cor"],
            "tipo": radar["tipo"],
        })

    # ordenar pior → melhor
    radar_lista.sort(key=lambda x: x["score"])

    context = {
        "radar_saude": radar_lista
    }

    return render(request, "cdv_api/radar_saude.html", context)

    # ordenar pior → melhor
    radar_lista.sort(key=lambda x: x["score"])

    context = {
        "radar_saude": radar_lista
    }

    return render(request, "cdv_api/radar_saude.html", context)


# =========================
# AÇÕES / APIs
# =========================

@login_required
def salvar_dados_cdv(request):
    key = "cdv_last_post"
    now = timezone.now()
    last = request.session.get(key)

    if last and (now - datetime.datetime.fromisoformat(last)).total_seconds() < 1.5:
        return JsonResponse(
            {"status": "error", "message": "Requisição ignorada (duplicada)."},
            status=429,
        )

    request.session[key] = now.isoformat()

    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Método não permitido."},
            status=405,
        )

    try:
        data = json.loads(request.body.decode("utf-8"))
        estacao_nome = data.get("estacao")
        transmissores_data = data.get("transmissores", [])
        receptores_data = data.get("receptores", [])

        if not estacao_nome:
            return JsonResponse(
                {"status": "error", "message": "Nome da estação não fornecido."},
                status=400,
            )

        try:
            estacao = Estacao.objects.get(nome=estacao_nome)
        except Estacao.DoesNotExist:
            return JsonResponse(
                {"status": "error", "message": f'Estação "{estacao_nome}" não encontrada.'},
                status=404,
            )

        hoje = timezone.localdate()

        with transaction.atomic():
            # ---------- TX ----------
            for tx in transmissores_data:
                num_circ = tx.get("num_circuito")
                num_tx = tx.get("num_transmissor")
                hora = tx.get("horario_coleta")
                temp = _pick_temp(tx)

                qs = (
                    Transmissor.objects.filter(
                        estacao=estacao,
                        num_circuito=num_circ,
                        num_transmissor=num_tx,
                        horario_coleta=hora,
                    )
                    .annotate(dia=TruncDate("data_manutencao"))
                    .filter(dia=hoje)
                )

                if qs.exists():
                    obj = qs.latest("id")
                    if temp is not None:
                        obj.temp_celsius = temp
                    obj.vout = safe_float(tx.get("vout"))
                    obj.pout = safe_float(tx.get("pout"))
                    obj.tap = tx.get("tap")
                    obj.tipo_transmissor = tx.get("tipo_transmissor")
                    obj.tipo_manutencao = _norm_manutencao(tx.get("tipo_manutencao"))
                    obj.save()
                else:
                    Transmissor.objects.create(
                        estacao=estacao,
                        num_circuito=num_circ,
                        num_transmissor=num_tx,
                        vout=safe_float(tx.get("vout")),
                        pout=safe_float(tx.get("pout")),
                        tap=tx.get("tap"),
                        tipo_transmissor=tx.get("tipo_transmissor"),
                        tipo_manutencao=_norm_manutencao(tx.get("tipo_manutencao")),
                        horario_coleta=hora,
                        temp_celsius=temp,
                    )

            # ---------- RX ----------
            for rx in receptores_data:
                num_circ = rx.get("num_circuito")
                num_rx = rx.get("num_receptor")
                hora = rx.get("horario_coleta")
                temp = _pick_temp(rx)

                iav = safe_float(rx.get("iav"))
                ith = safe_float(rx.get("ith"))

                if iav and ith and iav != 0:
                    rel_str = f"{(ith / iav) * 100:.2f}%"
                else:
                    rel = safe_float(rx.get("relacao"))
                    rel_str = f"{rel:.2f}%" if rel is not None else None

                qs = (
                    Receptor.objects.filter(
                        estacao=estacao,
                        num_circuito=num_circ,
                        num_receptor=num_rx,
                        horario_coleta=hora,
                    )
                    .annotate(dia=TruncDate("data_manutencao"))
                    .filter(dia=hoje)
                )

                if qs.exists():
                    obj = qs.latest("id")
                    if temp is not None:
                        obj.temp_celsius = temp
                    obj.iav = iav
                    obj.ith = ith
                    obj.relacao = rel_str
                    obj.tipo_manutencao = _norm_manutencao(rx.get("tipo_manutencao"))
                    obj.save()
                else:
                    Receptor.objects.create(
                        estacao=estacao,
                        num_circuito=num_circ,
                        num_receptor=num_rx,
                        iav=iav,
                        ith=ith,
                        relacao=rel_str,
                        tipo_manutencao=_norm_manutencao(rx.get("tipo_manutencao")),
                        horario_coleta=hora,
                        temp_celsius=temp,
                    )

        return JsonResponse({"status": "success", "message": "Dados salvos com sucesso!"})

    except Exception as e:
        logger.exception("Erro inesperado na view salvar_dados_cdv")
        return JsonResponse(
            {"status": "error", "message": f"Ocorreu um erro inesperado no servidor: {str(e)}"},
            status=500,
        )


@login_required
def listar_rxs_circuito(request):
    circuito = request.GET.get("circuito")
    estacao_id = request.GET.get("estacao_id")

    if not circuito:
        return JsonResponse({"erro": "Circuito não informado"}, status=400)

    receptores = Receptor.objects.filter(num_circuito=circuito)

    if estacao_id:
        receptores = receptores.filter(estacao_id=estacao_id)

    rxs = (
        receptores.values_list("num_receptor", flat=True)
        .distinct()
        .order_by("num_receptor")
    )

    rxs = [rx for rx in rxs if rx is not None]

    return JsonResponse({
        "circuito": circuito,
        "rxs": list(rxs),
    })


@login_required
def historico_circuito(request):
    circuito = request.GET.get("circuito")
    rx = request.GET.get("rx")
    estacao_id = request.GET.get("estacao_id")

    if not circuito:
        return JsonResponse({"erro": "Circuito não informado"}, status=400)

    receptores = Receptor.objects.filter(num_circuito=circuito)

    if estacao_id:
        receptores = receptores.filter(estacao_id=estacao_id)

    rx_num = None
    if rx:
        try:
            rx_num = int(str(rx).strip())
            receptores = receptores.filter(num_receptor=rx_num)
        except ValueError:
            return JsonResponse({"erro": "RX inválido"}, status=400)

    receptores = receptores.order_by("-data_manutencao", "-id")[:15]
    receptores = list(reversed(receptores))

    datas = []
    relacoes = []
    temperaturas = []

    for r in receptores:
        valor_relacao = relacao_para_float(r.relacao)
        if valor_relacao is None:
            continue

        data_txt = r.data_manutencao.strftime("%d/%m/%Y")
        hora_txt = r.horario_coleta.strftime("%H:%M") if r.horario_coleta else "--:--"

        datas.append(f"{data_txt} {hora_txt}")
        relacoes.append(valor_relacao)
        temperaturas.append(float(r.temp_celsius) if r.temp_celsius is not None else None)

    return JsonResponse({
        "datas": datas,
        "relacoes": relacoes,
        "temperaturas": temperaturas,
        "circuito": circuito,
        "rx": rx_num,
    })


# =========================
# EXCEL
# =========================

@login_required
def gerar_excel_estacao(request):
    estacao_id = request.GET.get("estacao_id")
    circuito_filtro = request.GET.get("circuito_filtro")
    data_inicio_str = request.GET.get("data_inicio")
    data_fim_str = request.GET.get("data_fim")
    tipo_manutencao_raw = request.GET.get("tipo_manutencao")

    def norm_tipo(val):
        if not val:
            return None
        v = val.strip().lower()
        if v.startswith("prevent"):
            return "preventiva"
        if v.startswith("corret"):
            return "corretiva"
        if v.startswith("check"):
            return "checklist"
        return None

    def _parse_date(s):
        try:
            return timezone.datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    tipo_manutencao = norm_tipo(tipo_manutencao_raw)

    if estacao_id:
        estacao = get_object_or_404(Estacao, id=estacao_id)
        transmissores = Transmissor.objects.filter(estacao=estacao)
        receptores = Receptor.objects.filter(estacao=estacao)
        nome_arquivo = f"dados_{estacao.nome}.xlsx"
    else:
        estacao = None
        transmissores = Transmissor.objects.all()
        receptores = Receptor.objects.all()
        nome_arquivo = "dados_todas_estacoes.xlsx"

    if circuito_filtro:
        transmissores = transmissores.filter(num_circuito__icontains=circuito_filtro)
        receptores = receptores.filter(num_circuito__icontains=circuito_filtro)

    if tipo_manutencao in {"preventiva", "corretiva", "checklist"}:
        transmissores = transmissores.filter(tipo_manutencao=tipo_manutencao)
        receptores = receptores.filter(tipo_manutencao=tipo_manutencao)

    data_inicio = _parse_date(data_inicio_str) if data_inicio_str else None
    data_fim = _parse_date(data_fim_str) if data_fim_str else None

    if data_inicio and data_fim:
        data_fim_ajustada = data_fim + timezone.timedelta(days=1)
        transmissores = transmissores.filter(
            data_manutencao__date__gte=data_inicio,
            data_manutencao__date__lt=data_fim_ajustada,
        )
        receptores = receptores.filter(
            data_manutencao__date__gte=data_inicio,
            data_manutencao__date__lt=data_fim_ajustada,
        )
    elif data_inicio:
        transmissores = transmissores.filter(data_manutencao__date__gte=data_inicio)
        receptores = receptores.filter(data_manutencao__date__gte=data_inicio)
    elif data_fim:
        data_fim_ajustada = data_fim + timezone.timedelta(days=1)
        transmissores = transmissores.filter(data_manutencao__date__lt=data_fim_ajustada)
        receptores = receptores.filter(data_manutencao__date__lt=data_fim_ajustada)

    transmissores = transmissores.exclude(temp_celsius__isnull=True)
    receptores = receptores.exclude(temp_celsius__isnull=True)

    transmissores = transmissores.order_by("estacao__nome", "data_manutencao", "horario_coleta", "id")
    receptores = receptores.order_by("estacao__nome", "data_manutencao", "horario_coleta", "id")

    wb = openpyxl.Workbook()

    # TX
    ws_tx = wb.active
    ws_tx.title = "Transmissores"
    ws_tx.append([
        "Estação", "Circuito", "TX", "VOUT", "POUT", "TAP", "Tipo TX",
        "Tipo Manutenção", "Data", "Horário Coleta", "Temp. (Celsius)"
    ])

    for t in transmissores:
        dt = timezone.localtime(t.data_manutencao) if t.data_manutencao else None
        data_fmt = dt.strftime("%d/%m/%Y") if dt else "-"
        hora_fmt = t.horario_coleta.strftime("%H:%M") if t.horario_coleta else "-"

        ws_tx.append([
            t.estacao.nome if t.estacao else "-",
            t.num_circuito,
            safe_int(t.num_transmissor),
            t.vout,
            t.pout,
            safe_int(t.tap),
            t.tipo_transmissor,
            t.tipo_manutencao,
            data_fmt,
            hora_fmt,
            t.temp_celsius,
        ])

    # RX
    ws_rx = wb.create_sheet("Receptores")
    ws_rx.append([
        "Estação", "Circuito", "RX", "IAV", "ITH", "Relação", "Tipo Manutenção",
        "Data", "Horário Coleta", "Temp. (Celsius)"
    ])

    linha_inicio_rx = ws_rx.max_row + 1

    for r in receptores:
        dt = timezone.localtime(r.data_manutencao) if r.data_manutencao else None
        data_fmt = dt.strftime("%d/%m/%Y") if dt else "-"
        hora_fmt = r.horario_coleta.strftime("%H:%M") if r.horario_coleta else "-"

        rel_excel = None
        if r.relacao and str(r.relacao).replace("%", "").strip():
            try:
                rel_excel = float(str(r.relacao).replace("%", "")) / 100.0
            except ValueError:
                rel_excel = None

        ws_rx.append([
            r.estacao.nome if r.estacao else "-",
            r.num_circuito,
            safe_int(r.num_receptor),
            r.iav,
            r.ith,
            rel_excel,
            r.tipo_manutencao,
            data_fmt,
            hora_fmt,
            r.temp_celsius,
        ])

    col_rel = get_column_letter(6)
    linha_fim_rx = ws_rx.max_row

    for row in range(linha_inicio_rx, linha_fim_rx + 1):
        ws_rx[f"{col_rel}{row}"].number_format = "0.00%"

    for ws in (ws_tx, ws_rx):
        last_col = ws.max_column
        for col_cells in ws.iter_cols(min_col=last_col, max_col=last_col, min_row=2, max_row=ws.max_row):
            for c in col_cells:
                c.number_format = "0.0"

    for ws in (ws_tx, ws_rx):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                v = cell.value
                l = len(str(v)) if v is not None else 0
                max_len = max(max_len, l)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    if linha_fim_rx >= linha_inicio_rx:
        intervalo = f"{col_rel}{linha_inicio_rx}:{col_rel}{linha_fim_rx}"
        fill_red = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        font_red = Font(color="FF9C0006")
        fill_green = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        font_green = Font(color="FF006100")

        ws_rx.conditional_formatting.add(
            intervalo,
            CellIsRule(operator="lessThan", formula=["0.6"], fill=fill_red, font=font_red)
        )
        ws_rx.conditional_formatting.add(
            intervalo,
            CellIsRule(operator="greaterThan", formula=["0.8"], fill=fill_red, font=font_red)
        )
        ws_rx.conditional_formatting.add(
            intervalo,
            FormulaRule(
                formula=[f"AND({col_rel}{linha_inicio_rx}>=0.6,{col_rel}{linha_inicio_rx}<=0.8)"],
                fill=fill_green,
                font=font_green,
            )
        )

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    wb.save(resp)
    return resp


# =========================
# DASHBOARD
# =========================

@login_required
def dashboard_manutencao(request):
    estacao_id = request.GET.get("estacao_id")
    circuito_filtro = request.GET.get("circuito_filtro")
    data_inicio = request.GET.get("data_inicio")
    data_fim = request.GET.get("data_fim")
    tipo_manutencao = request.GET.get("tipo_manutencao")

    lista_de_estacoes = ordenar_estacoes_linha(Estacao.objects.all())

    transmissores = Transmissor.objects.all()
    receptores = Receptor.objects.all()

    estacao_nome = None

    # FILTROS
    if estacao_id:
        estacao = get_object_or_404(Estacao, id=estacao_id)
        estacao_nome = estacao.nome
        transmissores = transmissores.filter(estacao=estacao)
        receptores = receptores.filter(estacao=estacao)

    if circuito_filtro:
        transmissores = transmissores.filter(num_circuito__icontains=circuito_filtro)
        receptores = receptores.filter(num_circuito__icontains=circuito_filtro)

    if tipo_manutencao:
        transmissores = transmissores.filter(tipo_manutencao=tipo_manutencao)
        receptores = receptores.filter(tipo_manutencao=tipo_manutencao)

    if data_inicio:
        transmissores = transmissores.filter(data_manutencao__date__gte=data_inicio)
        receptores = receptores.filter(data_manutencao__date__gte=data_inicio)

    if data_fim:
        transmissores = transmissores.filter(data_manutencao__date__lte=data_fim)
        receptores = receptores.filter(data_manutencao__date__lte=data_fim)

    # TOTAIS
    total_tx = transmissores.count()
    total_rx = receptores.count()

    tx_por_circuito = (
        transmissores.values("num_circuito")
        .annotate(total=Count("id"))
        .order_by("num_circuito")
    )

    rx_por_circuito = (
        receptores.values("num_circuito")
        .annotate(total=Count("id"))
        .order_by("num_circuito")
    )

    # TIPOS DE MANUTENÇÃO (TX + RX)
    contagem_tipos = {
        "preventiva": 0,
        "corretiva": 0,
        "checklist": 0,
    }

    for item in transmissores.values("tipo_manutencao").annotate(total=Count("id")):
        tipo = (item["tipo_manutencao"] or "").strip().lower()
        if tipo in contagem_tipos:
            contagem_tipos[tipo] += item["total"]

    for item in receptores.values("tipo_manutencao").annotate(total=Count("id")):
        tipo = (item["tipo_manutencao"] or "").strip().lower()
        if tipo in contagem_tipos:
            contagem_tipos[tipo] += item["total"]

    tipo_labels = []
    tipo_data = []

    for tipo, total in contagem_tipos.items():
        if total > 0:
            tipo_labels.append(tipo)
            tipo_data.append(total)

    # ÚLTIMOS RX / TX
    ultimos_tx = transmissores.order_by("-data_manutencao", "-id")[:10]

    ultimos_rx_qs = receptores.order_by("-data_manutencao", "-id")[:10]
    ultimos_rx = []

    for item in ultimos_rx_qs:
        valor_relacao = relacao_para_float(item.relacao)

        if valor_relacao is None:
            classe_relacao = ""
        elif valor_relacao < 60:
            classe_relacao = "relacao-baixa"
        elif valor_relacao > 80:
            classe_relacao = "relacao-alta"
        else:
            classe_relacao = "relacao-normal"

        ultimos_rx.append({
            "obj": item,
            "classe_relacao": classe_relacao,
        })

    # CONTAGEM POR CIRCUITO (PIOR RX)
    relacoes_por_circuito = {}
    contagem_abaixo_60 = 0
    contagem_entre_60_80 = 0
    contagem_acima_80 = 0

    receptores_ordenados = receptores.order_by("num_circuito", "-data_manutencao", "-id")
    agrupamento_circuitos = defaultdict(list)

    for r in receptores_ordenados:
        circuito = (r.num_circuito or "").strip().upper()
        valor = relacao_para_float(r.relacao)

        if valor is not None:
            agrupamento_circuitos[circuito].append({
                "obj": r,
                "relacao": valor,
            })

    for circuito, itens in agrupamento_circuitos.items():
        if not itens:
            continue

        abaixo_60 = [x for x in itens if x["relacao"] < 60]
        acima_80 = [x for x in itens if x["relacao"] > 80]

        if abaixo_60:
            pior_item = min(abaixo_60, key=lambda x: x["relacao"])
        elif acima_80:
            pior_item = max(acima_80, key=lambda x: x["relacao"])
        else:
            pior_item = min(itens, key=lambda x: x["relacao"])

        pior_relacao = pior_item["relacao"]
        classificacao, classe_relacao = classificar_relacao(pior_relacao)
        via = identificar_via(circuito)

        if pior_relacao < 60:
            contagem_abaixo_60 += 1
        elif pior_relacao > 80:
            contagem_acima_80 += 1
        else:
            contagem_entre_60_80 += 1

        relacoes_por_circuito[circuito] = {
            "circuito": circuito,
            "via": via,
            "rx_critico": pior_item["obj"].num_receptor,
            "relacao": round(pior_relacao, 2),
            "classificacao": classificacao,
            "classe_relacao": classe_relacao,
        }

    lista_relacoes = list(relacoes_por_circuito.values())
    lista_relacoes.sort(key=lambda x: (x["via"], x["circuito"]))

    # GRÁFICOS POR VIA (RX)
    relacao_labels_v1 = []
    relacao_data_v1 = []
    relacao_cores_v1 = []
    relacao_circuitos_v1 = []
    relacao_rxs_v1 = []

    relacao_labels_v2 = []
    relacao_data_v2 = []
    relacao_cores_v2 = []
    relacao_circuitos_v2 = []
    relacao_rxs_v2 = []

    receptores_grafico = receptores.order_by("num_circuito", "num_receptor", "-data_manutencao", "-id")
    ultimo_rx_por_circuito = {}

    for r in receptores_grafico:
        circuito = (r.num_circuito or "").strip().upper()
        rx = r.num_receptor
        chave = (circuito, rx)

        if chave in ultimo_rx_por_circuito:
            continue

        valor = relacao_para_float(r.relacao)
        if valor is None:
            continue

        ultimo_rx_por_circuito[chave] = {
            "circuito": circuito,
            "rx": rx,
            "relacao": round(valor, 2),
        }

    for item in ultimo_rx_por_circuito.values():
        label = [item["circuito"], f'RX {item["rx"]}']
        valor = item["relacao"]

        if valor < 60:
            cor = "rgba(255, 193, 7, 0.9)"
        elif valor <= 80:
            cor = "rgba(13, 110, 253, 0.9)"
        else:
            cor = "rgba(220, 53, 69, 0.9)"

        if item["circuito"].startswith("1"):
            relacao_labels_v1.append(label)
            relacao_data_v1.append(valor)
            relacao_cores_v1.append(cor)
            relacao_circuitos_v1.append(item["circuito"])
            relacao_rxs_v1.append(item["rx"])

        elif item["circuito"].startswith("2"):
            relacao_labels_v2.append(label)
            relacao_data_v2.append(valor)
            relacao_cores_v2.append(cor)
            relacao_circuitos_v2.append(item["circuito"])
            relacao_rxs_v2.append(item["rx"])

    # DEGRADAÇÃO GRADUAL
    circuitos_em_degradacao = detectar_degradacao_faixa(receptores)

    total_degradacao_negativa = sum(
        1 for item in circuitos_em_degradacao
        if item["tipo_degradacao"] == "Negativa"
    )

    total_degradacao_positiva = sum(
        1 for item in circuitos_em_degradacao
        if item["tipo_degradacao"] == "Positiva"
    )

    # MAPA DAS ESTAÇÕES
    estacoes_mapa = []
    todas_estacoes = ordenar_estacoes_linha(Estacao.objects.all())

    for est in todas_estacoes:
        rx_est = Receptor.objects.filter(estacao=est)

        if data_inicio:
            rx_est = rx_est.filter(data_manutencao__date__gte=data_inicio)

        if data_fim:
            rx_est = rx_est.filter(data_manutencao__date__lte=data_fim)

        if tipo_manutencao:
            rx_est = rx_est.filter(tipo_manutencao=tipo_manutencao)

        if circuito_filtro:
            rx_est = rx_est.filter(num_circuito__icontains=circuito_filtro)

        relacoes_est = []
        for r in rx_est.order_by("-data_manutencao", "-id"):
            valor = relacao_para_float(r.relacao)
            if valor is not None:
                relacoes_est.append(valor)

        degradacoes_est = detectar_degradacao_faixa(rx_est)
        qtd_criticos = sum(1 for v in relacoes_est if v < 60 or v > 80)

        if qtd_criticos > 0:
            status = "critico"
        elif len(degradacoes_est) > 0:
            status = "atencao"
        else:
            status = "normal"

        estacoes_mapa.append({
            "id": est.id,
            "nome": est.nome,
            "status": status,
            "qtd_criticos": qtd_criticos,
            "qtd_degradacoes": len(degradacoes_est),
        })

# GRÁFICO DE TENDÊNCIA DE DEGRADAÇÃO
    degradacao_datasets = []

    for item in circuitos_em_degradacao:
        circuito = (item.get("circuito") or "").strip()

        qs = (
            receptores
            .filter(num_circuito__iexact=circuito)
            .order_by("data_manutencao", "horario_coleta", "id")
        )

        labels = []
        values = []

        for r in qs:
            valor = relacao_para_float(r.relacao)
            if valor is None:
                continue

            if r.horario_coleta:
                tempo = f"{r.data_manutencao.strftime('%d/%m/%Y')} {r.horario_coleta.strftime('%H:%M')}"
            else:
                tempo = r.data_manutencao.strftime('%d/%m/%Y')

            labels.append(tempo)
            values.append(round(valor, 2))

        if not values:
            continue

        negativa = item["tipo_degradacao"] == "Negativa"

        degradacao_datasets.append({
            "label": f"Circuito {circuito}",
            "labels": labels,
            "values": values,
            "borderColor": "rgba(255, 193, 7, 1)" if negativa else "rgba(220, 53, 69, 1)",
            "backgroundColor": "rgba(255, 193, 7, 0.15)" if negativa else "rgba(220, 53, 69, 0.15)",
            "borderDash": [8, 6] if negativa else [],
            "pointRadius": 3,
            "pointHoverRadius": 5,
            "tension": 0.35,
        })
    
    context = {
        "lista_de_estacoes": lista_de_estacoes,
        "selected_estacao_id": str(estacao_id) if estacao_id else "",
        "estacao_nome": estacao_nome,
        "estacoes_mapa": estacoes_mapa,

        "circuito_filtro": circuito_filtro or "",
        "tipo_manutencao": tipo_manutencao or "",
        "data_inicio": data_inicio or "",
        "data_fim": data_fim or "",

        "total_tx": total_tx,
        "total_rx": total_rx,

        "tx_labels": json.dumps([x["num_circuito"] for x in tx_por_circuito]),
        "tx_data": json.dumps([x["total"] for x in tx_por_circuito]),

        "rx_labels": json.dumps([x["num_circuito"] for x in rx_por_circuito]),
        "rx_data": json.dumps([x["total"] for x in rx_por_circuito]),

        "tipo_labels": json.dumps(tipo_labels),
        "tipo_data": json.dumps(tipo_data),

        "relacao_labels_v1": json.dumps(relacao_labels_v1),
        "relacao_data_v1": json.dumps(relacao_data_v1),
        "relacao_cores_v1": json.dumps(relacao_cores_v1),
        "relacao_circuitos_v1": json.dumps(relacao_circuitos_v1),
        "relacao_rxs_v1": json.dumps(relacao_rxs_v1),

        "relacao_labels_v2": json.dumps(relacao_labels_v2),
        "relacao_data_v2": json.dumps(relacao_data_v2),
        "relacao_cores_v2": json.dumps(relacao_cores_v2),
        "relacao_circuitos_v2": json.dumps(relacao_circuitos_v2),
        "relacao_rxs_v2": json.dumps(relacao_rxs_v2),

        "contagem_abaixo_60": contagem_abaixo_60,
        "contagem_entre_60_80": contagem_entre_60_80,
        "contagem_acima_80": contagem_acima_80,

        "circuitos_em_degradacao": circuitos_em_degradacao,
        "total_degradacao_gradual": len(circuitos_em_degradacao),
        "total_degradacao_negativa": total_degradacao_negativa,
        "total_degradacao_positiva": total_degradacao_positiva,

        "degradacao_datasets": json.dumps(degradacao_datasets),
        
        "ultimos_tx": ultimos_tx,
        "ultimos_rx": ultimos_rx,
    }

    return render(request, "cdv_api/dashboard_manutencao.html", context)